import os
import glob
from openpyxl import load_workbook
import pandas as pd

def clean_name(name):
    if not name: return ""
    return str(name).lower().replace('?', '').replace(' ', '').strip()

photos_dir = r'C:\Users\Admin\Desktop\CLASS WISE\Photos\Class 1'
excel_path = r'C:\Users\Admin\Desktop\CLASS WISE\I-A.xlsx'

photos = glob.glob(os.path.join(photos_dir, '*.*'))
photo_map = {}
for p in photos:
    basename = os.path.basename(p)
    name_part = os.path.splitext(basename)[0]
    cleaned = clean_name(name_part)
    photo_map[cleaned] = p

wb = load_workbook(excel_path)
ws = wb.active

header_row = 1
name_col = None
for col in range(1, ws.max_column + 1):
    val = str(ws.cell(row=header_row, column=col).value).strip().lower()
    if 'name' in val and 'father' not in val and 'mother' not in val:
        name_col = col

used_photos = set()
unmatched_excel_names = []

for row_idx in range(header_row + 1, ws.max_row + 1):
    name_cell = ws.cell(row=row_idx, column=name_col).value
    cleaned_excel_name = clean_name(name_cell)
    
    if not cleaned_excel_name: continue
        
    matched_photo = photo_map.get(cleaned_excel_name)
    if not matched_photo:
        for p_name, p_path in photo_map.items():
            if cleaned_excel_name in p_name or p_name in cleaned_excel_name:
                matched_photo = p_path
                break
                
    if matched_photo:
        used_photos.add(matched_photo)
    else:
        unmatched_excel_names.append(name_cell)

unmatched_photos = []
for p in photos:
    if p not in used_photos:
        unmatched_photos.append(os.path.basename(p))

with open('match_results.txt', 'w', encoding='utf-8') as f:
    f.write("Unmatched Photos:\n")
    for p in unmatched_photos:
        f.write(f"- {p}\n")
    f.write("\nUnmatched Excel Names:\n")
    for n in unmatched_excel_names:
        f.write(f"- {n}\n")
