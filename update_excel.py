import sqlite3
import pandas as pd
import json
import base64
import io
import os
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

db_path = r'C:\Users\Admin\AppData\Roaming\school-management-system\school.db'
excel_path = r'C:\Users\Admin\Desktop\Annual Exam 2026\TPS DATA EXCEL 28.07.xls'
out_excel = r'C:\Users\Admin\Desktop\Annual Exam 2026\TPS DATA EXCEL 28.07_HighQuality.xlsx'

print("Loading DB...")
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
rows = conn.execute("SELECT * FROM students").fetchall()

db_students = [dict(r) for r in rows]

df = pd.read_excel(excel_path, engine='xlrd')

existing_admission_nos = set(df['Admission No.'].dropna().astype(str).str.replace(r'\.0$', '', regex=True).str.strip().str.lower())
existing_names = set(df['Full Name'].dropna().astype(str).str.strip().str.lower())

missing_students = []
for s in db_students:
    adm_no = str(s.get('admissionNo', '')).replace('.0', '').strip().lower()
    name = str(s.get('fullName', '')).strip().lower()
    if adm_no and adm_no in existing_admission_nos: continue
    if not adm_no and name in existing_names: continue
    missing_students.append(s)

new_rows = []
for s in missing_students:
    new_rows.append({
        'Roll No.': s.get('rollNo', ''),
        'Full Name': s.get('fullName', ''),
        'PHOTO NO.': '',
        'Class': s.get('className', ''),
        'Admission No.': s.get('admissionNo', ''),
        'Date of Birth': s.get('dob', ''),
        'Father Name': s.get('fatherName', ''),
        'Mother Name': s.get('motherName', ''),
        'Phone': s.get('phone', ''),
        'Address': s.get('address', '')
    })

if new_rows:
    df_missing = pd.DataFrame(new_rows)
    df = pd.concat([df, df_missing], ignore_index=True)

df.to_excel(out_excel, index=False)

wb = load_workbook(out_excel)
ws = wb.active

max_col = ws.max_column
photo_col_idx = max_col + 1
ws.cell(row=1, column=photo_col_idx, value='Photo Image')

font_style = Font(size=14, bold=True)
center_align = Alignment(vertical='center', horizontal='center', wrap_text=True)

for col_idx in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 25

ws.column_dimensions[get_column_letter(photo_col_idx)].width = 35

for row in ws.iter_rows():
    for cell in row:
        cell.font = font_style
        cell.alignment = center_align

ws.row_dimensions[1].height = 40

created_temps = []
for row_idx in range(2, ws.max_row + 1):
    adm_cell = ws.cell(row=row_idx, column=5).value 
    name_cell = ws.cell(row=row_idx, column=2).value
    
    adm = str(adm_cell).replace('.0', '').strip().lower() if pd.notna(adm_cell) and adm_cell else ""
    name = str(name_cell).strip().lower() if pd.notna(name_cell) and name_cell else ""
    
    matched = None
    if adm:
        for s in db_students:
            if str(s.get('admissionNo', '')).replace('.0', '').strip().lower() == adm:
                matched = s; break
    if not matched and name:
        for s in db_students:
            if str(s.get('fullName', '')).strip().lower() == name:
                matched = s; break
                
    ws.row_dimensions[row_idx].height = 60
    
    if matched and matched.get('photo'):
        photo_b64 = matched['photo']
        if photo_b64.startswith('data:image'): photo_b64 = photo_b64.split(',')[1]
        photo_b64 += '=' * (-len(photo_b64) % 4)
            
        try:
            image_data = base64.b64decode(photo_b64)
            img = PILImage.open(io.BytesIO(image_data))
            
            temp_path = f"temp_{row_idx}.jpg"
            if img.mode != 'RGB': img = img.convert('RGB')
            # Save physical file at 100% original quality and size
            img.save(temp_path, quality=100)
            created_temps.append(temp_path)
            
            xl_img = Image(temp_path)
            
            # Only resize VISUALLY in Excel, keep underlying physical pixels intact
            if xl_img.height > 0:
                aspect = xl_img.width / xl_img.height
                xl_img.height = 200
                xl_img.width = int(200 * aspect)
            
            ws.row_dimensions[row_idx].height = 160
            
            ws.add_image(xl_img, f"{get_column_letter(photo_col_idx)}{row_idx}")
            
        except Exception as e:
            print(f"Error inserting photo for row {row_idx}: {e}")

wb.save(out_excel)
for f in created_temps:
    try: os.remove(f)
    except: pass
print("Done! Saved to:", out_excel)
