import os
import glob
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import re

def clean_name(name):
    if not name: return ""
    return re.sub(r'[^a-z0-9]', '', str(name).lower())

photos_dir = r'C:\Users\Admin\Desktop\CLASS WISE\Photos\Class 1'
excel_path = r'C:\Users\Admin\Desktop\CLASS WISE\I-A.xlsx'
out_excel = r'C:\Users\Admin\Desktop\CLASS WISE\I-A_Updated_v2.xlsx'

photos = glob.glob(os.path.join(photos_dir, '*.*'))
photo_map = {}
for p in photos:
    basename = os.path.basename(p)
    name_part = os.path.splitext(basename)[0]
    cleaned = clean_name(name_part)
    photo_map[cleaned] = p

print(f"Found {len(photo_map)} photos.")

wb = load_workbook(excel_path)
ws = wb.active

header_row = 1
name_col = None
photo_col = None
for col in range(1, ws.max_column + 1):
    val = str(ws.cell(row=header_row, column=col).value).strip().lower()
    if 'name' in val and 'father' not in val and 'mother' not in val:
        name_col = col
    if 'photo' in val:
        photo_col = col

if not photo_col:
    photo_col = ws.max_column + 1
    ws.cell(row=header_row, column=photo_col, value='Photo')

if not name_col:
    print("Could not find Name column!")
    exit(1)

from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(photo_col)].width = 35

matched_count = 0
for row_idx in range(header_row + 1, ws.max_row + 1):
    name_cell = ws.cell(row=row_idx, column=name_col).value
    cleaned_excel_name = clean_name(name_cell)
    
    if not cleaned_excel_name:
        continue
        
    ws.row_dimensions[row_idx].height = 60
    
    # Try exact match
    matched_photo = photo_map.get(cleaned_excel_name)
    
    # Try partial match if no exact
    if not matched_photo:
        for p_name, p_path in photo_map.items():
            if cleaned_excel_name in p_name or p_name in cleaned_excel_name:
                matched_photo = p_path
                break
                
    if matched_photo:
        try:
            img = PILImage.open(matched_photo)
            
            temp_path = f"temp_class1_{row_idx}.jpg"
            if img.mode != 'RGB':
                img = img.convert('RGB')
            img.save(temp_path, quality=90)
            
            xl_img = Image(temp_path)
            
            if xl_img.height > 0:
                aspect = xl_img.width / xl_img.height
                xl_img.height = 200
                xl_img.width = int(200 * aspect)
            
            ws.row_dimensions[row_idx].height = 160
            ws.add_image(xl_img, f"{get_column_letter(photo_col)}{row_idx}")
            matched_count += 1
            
            # Remove used photo from map so it doesn't match again
            for k, v in list(photo_map.items()):
                if v == matched_photo:
                    del photo_map[k]
                    break
            
        except Exception as e:
            print(f"Error for row {row_idx}: {e}")

wb.save(out_excel)
print(f"Done! Matched {matched_count} photos. Saved to {out_excel}")

for f in glob.glob("temp_class1_*.jpg"):
    try: os.remove(f)
    except: pass
